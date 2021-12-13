import pandas as pd
import os
import sys
import warnings
import openpyxl
import javalang
import random
from multiprocessing import Process, Queue
warnings.filterwarnings('ignore')
sys.setrecursionlimit(50000)

class Pipeline:
    def __init__(self,  ratio, root):
        self.ratio = ratio
        self.root = root
        self.project = None
        self.size = None

    def get_project_info(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb['star1000']
        max_row = ws.max_row
        name = []
        for i in range(1,max_row+1):
            name.append(ws[i][0].value.split('/')[-1])
        self.project = name

    def parse_source(self, project_name):
        print(project_name,self.root+project_name+'/'+project_name+'.pkl')
        path = self.root+project_name+'/'+project_name+'_ast.pkl'
        if os.path.exists(path):
            source = pd.read_pickle(path)
        else:
            def parse_program(func):
                tokens = javalang.tokenizer.tokenize(func)
                parser = javalang.parser.Parser(tokens)
                tree = parser.parse_member_declaration()
                return tree
            source = pd.read_pickle(self.root+project_name+'/'+project_name+'.pkl')
            source.columns = ['old', 'new', 'label']
            source['old'] = source['old'].apply(parse_program)
            source['new'] = source['new'].apply(parse_program)
            source.to_pickle(path)

    # split data for training, developing and testing
    def split_data(self, project_name):
        data_path = self.root + project_name +'/'
        data = pd.read_pickle(data_path + project_name + '_ast.pkl')
        data_num = len(data)
        ratios = [int(r) for r in self.ratio.split(':')]
        train_split = int(ratios[0]/sum(ratios)*data_num)
        val_split = train_split + int(ratios[1]/sum(ratios)*data_num)

        data = data.sample(frac=1, random_state=666)
        train = data.iloc[:train_split]
        dev = data.iloc[train_split:val_split]
        test = data.iloc[val_split:]

        def check_or_create(path):
            if not os.path.exists(path):
                os.mkdir(path)
        train_path = data_path+'train/'
        check_or_create(train_path)
        train_file_path = train_path+'train_.pkl'
        train.to_pickle(train_file_path)

        dev_path = data_path+'dev/'
        check_or_create(dev_path)
        dev_file_path = dev_path+'dev_.pkl'
        dev.to_pickle(dev_file_path)

        test_path = data_path+'test/'
        check_or_create(test_path)
        test_file_path = test_path+'test_.pkl'
        test.to_pickle(test_file_path)

    # construct dictionary and train word embedding
    def dictionary_and_embedding(self, project_name, size):
        self.size = size
        data_path = self.root + project_name + '/'
        input_file = data_path + 'train/train_.pkl'
        data = pd.read_pickle(input_file)
        if not os.path.exists(data_path+'train/embedding'):
            os.mkdir(data_path+'train/embedding')

        from utils import get_sequence as func

        def trans_to_sequences(ast):
            sequence = []
            func(ast, sequence)
            return sequence
        corpus = data['old'].apply(trans_to_sequences)
        corpus += data['new'].apply(trans_to_sequences) 
        str_corpus = [' '.join(c) for c in corpus]

        from gensim.models.word2vec import Word2Vec
        w2v = Word2Vec(corpus, size=size, sg=1, min_count = 3) # max_final_vocab=3000 tmp ignore
        w2v.save(data_path+'train/embedding/node_w2v_' + str(size))

    # generate block sequences with index representations
    def generate_block_seqs(self, project_name):
        from utils import get_blocks_v1 as func
        from gensim.models.word2vec import Word2Vec

        word2vec = Word2Vec.load(self.root+project_name+'/train/embedding/node_w2v_' + str(self.size)).wv
        vocab = word2vec.vocab
        max_token = word2vec.syn0.shape[0]

        def tree_to_index(node):
            token = node.token
            result = [vocab[token].index if token in vocab else max_token]
            children = node.children
            for child in children:
                result.append(tree_to_index(child))
            return result

        def trans2seq(r):
            blocks = []
            func(r, blocks)
            tree = []
            for b in blocks:
                btree = tree_to_index(b)
                tree.append(btree)
            return tree

        def reduce_line(a):
            return eval('['+str(a).replace(' ','').replace('[','').replace(']','')+']')

        path = self.root + project_name + '/'
        
        train = pd.read_pickle(path+'train/train_.pkl')
        train['old'] = train['old'].apply(trans2seq)
        train['new'] = train['new'].apply(trans2seq)
        dellist = []
        '''
        for _,item in test.iterrows():
            if len(reduce_line(item['old']))>1000 or len(reduce_line(item['new']))>1000 or item['old']==item['new']:
                dellist.append(_)
        '''
        one = []
        zero = []
        for _,item in train.iterrows():
            if len(reduce_line(item['old']))>1000 or len(reduce_line(item['new']))>1000 or item['old']==item['new']:
                dellist.append(_)
            else:
                if item['label']==0:zero.append(_)
                else:one.append(_)
        ls = min(len(one),len(zero))
        random.shuffle(one)
        random.shuffle(zero)
        #if len(one)!=ls:dellist += one[ls:]
        #else : dellist += zero[ls:]
        train.drop(dellist,inplace = True)
        train.to_pickle(path+'train/blocks.pkl')

        test = pd.read_pickle(path+'test/test_.pkl')
        test['old'] = test['old'].apply(trans2seq)
        test['new'] = test['new'].apply(trans2seq)
        dellist = []
        for _,item in test.iterrows():
            if len(reduce_line(item['old']))>1000 or len(reduce_line(item['new']))>1000 or item['old']==item['new']:
                dellist.append(_)
        test.drop(dellist,inplace = True)
        test.to_pickle(path+'test/blocks.pkl')

        dev = pd.read_pickle(path+'dev/dev_.pkl')
        dev['old'] = dev['old'].apply(trans2seq)
        dev['new'] = dev['new'].apply(trans2seq)
        dellist = []
        for _,item in dev.iterrows():
            if len(reduce_line(item['old']))>1000 or len(reduce_line(item['new']))>1000 or item['old']==item['new']:
                dellist.append(_)
        dev.drop(dellist,inplace = True)
        dev.to_pickle(path+'dev/blocks.pkl')

    # run for processing data to train
    def run(self):
        print('get project information...')
        #self.get_project_info('pj.xlsx')
        self.project = ['accumulo','ambari','cloudstack','commons-lang','flink']
        print('parse source code...')
        for name in self.project:
            self.parse_source(name)
        print('split data...')
        for name in self.project:
            self.split_data(name)
        print('train word embedding...')
        for name in self.project:
            self.dictionary_and_embedding(name,128) #后续可看256的效果
        print('generate block sequences...')
        for name in self.project:
            self.generate_block_seqs(name)


if __name__=='__main__':
    ppl = Pipeline('3:1:1', 'data/')
    ppl.run()
    