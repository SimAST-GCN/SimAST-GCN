import os
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader
import gensim
import random
import pandas as pd
import torch
import time,sys
import numpy as np
import warnings
from gensim.models.word2vec import Word2Vec
from ast_model import BatchProgramCC
from torch.autograd import Variable
from sklearn.metrics import precision_recall_fscore_support
from sklearn import metrics
from sklearn.model_selection import StratifiedKFold,KFold
from sklearn.utils.class_weight import compute_class_weight 
from openpyxl import Workbook,load_workbook
from sklearn.metrics import matthews_corrcoef
warnings.filterwarnings('ignore')
'''
class MyDataset(Dataset):
    def __init__(self, file_path):
        source = pd.read_pickle(file_path)
        self.len = len(source)
        self.old = source['old'].tolist()
        self.new = source['new'].tolist()
        self.label = source['label'].tolist()

    def __len__(self):
        return self.len

    def __getitem__(self, index):
        return self.old[index], self.new[index], self.label[index]
'''
os.environ['CUDA_VISIBLE_DEVICES']='0'

def get_batch(dataset, idx, bs):
    tmp = dataset.iloc[idx: idx+bs]
    x1, x2, labels = [], [], []
    for _, item in tmp.iterrows():
        x1.append(item['old'])
        x2.append(item['new'])
        labels.append([item['label']])
    return x1, x2, torch.FloatTensor(labels)

#if __name__ == '__main__':
def run_astnn(root):
    #parameter
    BATCH_SIZE = 100
    EPOCHS = 100
    learning_rate = 0.01
    HIDDEN_DIM = 100
    ENCODE_DIM = 128

    ott = []

    #root = 'data/ambari/'
    train_data = pd.read_pickle(root+'train/blocks.pkl').sample(frac=1)
    test_data = pd.read_pickle(root+'test/blocks.pkl').sample(frac=1)
    dev_data = pd.read_pickle(root + 'dev/blocks.pkl').sample(frac=1)

    word2vec = Word2Vec.load(root + 'train/embedding/node_w2v_128').wv
    MAX_TOKENS = word2vec.syn0.shape[0]
    EMBEDDING_DIM = word2vec.syn0.shape[1]
    embeddings = np.zeros((MAX_TOKENS + 1, EMBEDDING_DIM), dtype="float32")
    embeddings[:word2vec.syn0.shape[0]] = word2vec.syn0

    model = BatchProgramCC(EMBEDDING_DIM,HIDDEN_DIM,MAX_TOKENS+1,ENCODE_DIM,BATCH_SIZE, embeddings)
    model.cuda()

    # Loss and optimizer
    parameters = model.parameters()

    #compute_weight
    l = train_data['label'].tolist()
    class_weight = 'balanced'
    classes = np.array([0,1])
    weight = compute_class_weight(class_weight,classes,l)
    criterion = nn.CrossEntropyLoss(weight = torch.from_numpy(weight).cuda().float())
    print(weight)
    #criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.SGD(model.parameters(), lr=learning_rate, momentum=0.9)

    #train and dev
    tm = 0
    precision, recall, f1 = 0, 0, 0
    train_data_t, dev_data_t, test_data_t = train_data, dev_data, test_data
    for epoch in range(EPOCHS):
        print('Training : ',epoch,' project :',root)
        model.train()
        start_time = time.time()
        total_loss = 0.0
        i=0
        while i < len(train_data_t):
            batch = get_batch(train_data_t, i, BATCH_SIZE)
            i += BATCH_SIZE
            old, new, labels = batch
            labels = labels.cuda().squeeze()
            model.batch_size = len(labels)
            model.hidden = model.init_hidden()
            #print("in")
            outputs = model(old,new)
            #print("out")
            #print(outputs.cpu())
            #print(outputs.shape,labels.shape)
            loss = criterion(outputs,labels.long())
            #backward and optimize
            total_loss += loss.item()
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            if (i/BATCH_SIZE)%10==0:
                end_time = time.time()
                tm +=(end_time-start_time)
                print ('Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}, Time: {:.4f}' 
                    .format(epoch+1, EPOCHS, i/BATCH_SIZE, 1.0*len(train_data_t)/BATCH_SIZE, total_loss/10, end_time-start_time))
                start_time = time.time()
                total_loss = 0
        print('Testing')
        model.eval()

        lb = torch.Tensor()
        pr = torch.Tensor()
        correct = 0
        total = 0

        i=0
        while i < len(test_data_t):
            batch = get_batch(test_data_t, i, BATCH_SIZE)
            i += BATCH_SIZE
            old, new, labels = batch
            #if len(labels)==0:break
            labels = labels.cuda().squeeze()
            model.batch_size = len(labels)
            model.hidden = model.init_hidden()
            outputs = model(old,new)
            loss = criterion(outputs,labels.long())
            _,predicted = torch.max(outputs.data,1)
            total+=labels.size(0)
            pr = torch.cat((pr,predicted.cpu()),0)
            lb = torch.cat((lb,labels.cpu()),0)
            correct += (predicted == labels).sum().item()
            print('Test Accuracy of the model on the {} test case: {:.4f} %, loss: {:.4f}'.format(total,100 * correct / total, loss.item()))

        print(pr)
        print(lb)
        zero = 0
        zero_all = 0
        one = 0
        one_all = 0
        for i in range(len(lb)):
            if lb[i]==0:
                zero_all+=1
                if pr[i]==0:zero+=1
            else:
                one_all+=1
                if pr[i]==1:one+=1
        print("Test one acc: {}/{}, zero acc: {}/{}".format(one,one_all,zero,zero_all))
        print("Recall :",metrics.recall_score(lb,pr))
        print("F1 :",metrics.f1_score(lb,pr))
        print("AUC :",metrics.roc_auc_score(lb,pr))
        print("MCC : {}".format(matthews_corrcoef(lb,pr)))
        ott.append(['','',metrics.recall_score(lb,pr),metrics.f1_score(lb,pr),metrics.roc_auc_score(lb,pr),100*correct/total,matthews_corrcoef(lb,pr),tm])
    return ott


if __name__ == '__main__':
    p = sys.argv[1]
    #root = 'data/ambari/'
    project = 'data/'+p+'/'
    out = run_astnn(project)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'project'
    ws['B1'] = 'model'
    ws['C1'] = 'Recall'
    ws['D1'] = 'F1'
    ws['E1'] = 'AUC'
    ws['F1'] = 'ACCURACY'
    ws['G1'] = 'MCC'
    ws['H1'] = 'Time'

    file_path = './rerun/ASTNN_'+p+'.xlsx'

    out[0][0]=project
    out[0][1]='ASTNN'
    for row in out:
        ws.append(row)
    wb.save(file_path)

