import torch 
import torch.nn as nn
import torchvision
import torchvision.transforms as transforms
from torch.utils.data import Dataset, DataLoader
from gensim.models.word2vec import Word2Vec
import json
import random
import gensim
import numpy as np
import pandas as pd
import os,time,sys
from sklearn import metrics
import warnings
import torch.nn.utils.rnn as rnn_utils
from sklearn.utils.class_weight import compute_class_weight 
warnings.filterwarnings('ignore')
from openpyxl import Workbook,load_workbook
from sklearn.metrics import matthews_corrcoef

os.environ["CUDA_VISIBLE_DEVICES"] = "0"
# Device configuration
#device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
device = torch.device('cuda')

#Dataset
class MyDataset(Dataset):
    def __init__(self,file_path):
        labels = []
        old = []
        new = []
        source = pd.read_pickle(file_path)
        self.len = len(source)
        self.label = source['label'].tolist()
        self.old = source['old'].tolist()
        self.new = source['new'].tolist()

    def __len__(self):
        return self.len

    def __getitem__(self,index):
        return self.old[index],self.new[index],self.label[index]

#Module
class EncoderRNN(nn.Module):
    def __init__(self, input_size, hidden_size):
        super(EncoderRNN, self).__init__()
        self.hidden_size = hidden_size
        self.gru = nn.GRU(input_size, hidden_size)

    def forward(self, input):
        output, hidden = self.gru(input)
        return output, hidden

    def initHidden(self):
        return torch.zeros(1, 1, self.hidden_size, device=device)
    
class DecoderRNN(nn.Module):
    def __init__(self, input_size, hidden_size):
        super(DecoderRNN, self).__init__()
        self.hidden_size = hidden_size
        self.gru = nn.GRU(input_size, hidden_size)

    def forward(self, input):
        output, hidden = self.gru(input)
        return output, hidden

    def initHidden(self):
        return torch.zeros(1, 1, self.hidden_size, device=device)

class RCF(nn.Module):
    def __init__(self,weights,vocab_size):
        super(RCF,self).__init__()
        self.embedding_size = 128
        self.window_sizes = [2,3,4]
        self.feature_size = 100
        self.hidden_size = 300
        self.hidden_gru = 256
        self.num_classes = 2
        self.text_lin = 50  #code最多50行
        self.text_len = 100  #每行最多50个token
        self.num_layers = 1
        self.embedding=nn.Embedding(vocab_size,self.embedding_size)
        self.embedding.weight.data.copy_(torch.from_numpy(weights))
        self.convs = nn.ModuleList([
            nn.Sequential(nn.Conv1d(in_channels = self.embedding_size,out_channels = self.feature_size,kernel_size = h),
                          nn.BatchNorm1d(num_features=self.feature_size),
                          nn.ReLU(),
                          nn.MaxPool1d(kernel_size = self.text_len-h+1)
            )
            for h in self.window_sizes
        ])
        self.maxpool = nn.MaxPool1d(kernel_size = 2)  #fuse operation
        self.lstm = nn.LSTM(self.feature_size*len(self.window_sizes), self.hidden_size, self.num_layers, batch_first=True)        
        self.encoder = EncoderRNN(self.feature_size*len(self.window_sizes),self.hidden_gru)
        self.decoder = DecoderRNN(self.hidden_gru,self.feature_size*len(self.window_sizes))
        self.maxpool2 = nn.MaxPool1d(kernel_size = 30, stride = 30)
        self.fc = nn.Linear(in_features = 500,out_features = 256)
        self.fc2 = nn.Linear(in_features = 256,out_features = 100)
        self.fc3 = nn.Linear(in_features = 100,out_features = self.num_classes)

    
    def forward(self, old, new):
        #print(old.shape)
        emb_old = self.Emb(old)  #[32, 50, 50, 128]
        emb_new = self.Emb(new)  
        con_old = self.Conv(emb_old)  #[32, 50, 300]
        con_new = self.Conv(emb_new)
        lstm_old = self.LSTM(con_old)  #[32, 50, 300]
        lstm_new = self.LSTM(con_new)
        lstm_all = torch.stack((lstm_old,lstm_new),0)  #[2, 32, 50, 300]
        fuse = self.Fuse(lstm_all)  #[32, 50, 300]
        de_encode = self.Decode_Encode(fuse)  #[32, 50, 300]
        out = self.maxpool2(de_encode)  #[32, 50, 10]
        out = self.fc(out.reshape(out.size(0),-1))  #[32, 500] -> [32, 256]
        out = self.fc2(out)  #[32, 100]
        out = self.fc3(out)  #[32, 2]
        return out

    def Emb(self,idx):
        idx = idx.permute(1,0,2)
        out = torch.Tensor().to(device)
        for _ in idx:
            out = torch.cat((out,self.embedding(_)),0)
        return out.reshape(self.text_lin,-1,self.text_len,self.embedding_size).permute(1,0,2,3)

    def Conv(self,embed):
        out = torch.Tensor().to(device)
        embed = embed.permute(1,0,3,2)
        for _ in embed:
            tmp = [conv(_) for conv in self.convs]
            #print(tmp.shape)
            tmp = torch.cat(tmp,dim=1)
            tmp = tmp.view(-1,tmp.size(1))  
            out = torch.cat((out,tmp),0)
        return out.reshape(self.text_lin,-1,out.size(1)).permute(1,0,2)
    
    def LSTM(self,con):
        con = con.permute(1,0,2)
        out,_ = self.lstm(con)
        return out.permute(1,0,2)
        '''
        con = con.permute(1,0,2)
        for _ in con:
            tmp,(h,c) = self.lstm(_)
            out = torch.cat((out,tmp),0)
        return out.reshape(self.text_lin,-1,out.size(1)).permute(1,0,2)
        '''

    def Decode_Encode(self,fuse):
        fuse = fuse.permute(1,0,2)
        out,_ = self.encoder(fuse)
        out,_ = self.decoder(out)
        return out.permute(1,0,2)

    def Fuse(self,comb):
        out = torch.Tensor().to(device)
        comb = comb.permute(2,1,3,0)
        for _ in comb:
            tmp = self.maxpool(_)
            tmp = tmp.view(-1,tmp.size(1))
            out = torch.cat((out,tmp),0)
        return out.reshape(self.text_lin,-1,out.size(1)).permute(1,0,2)

    def initHidden(self):
        return torch.zeros(1,1,self.hidden_size,device=device)

def nd(p):
    train_path = './data/'+p+'/token/'+p+'_token_train.pkl'
    test_path = './data/'+p+'/token/'+p+'_token_test.pkl'
    valid_path = './data/'+p+'/token/'+p+'_token_dev.pkl'
    train_dataset = MyDataset(train_path)
    test_dataset = MyDataset(test_path)
    valid_dataset = MyDataset(valid_path)

    batch_size = 256

    train_loader = DataLoader(dataset = train_dataset,batch_size = batch_size,shuffle = True)
    test_loader = DataLoader(dataset = test_dataset,batch_size = batch_size,shuffle = False)
    valid_loader = DataLoader(dataset = valid_dataset,batch_size = batch_size,shuffle = False)

    input_size = 128
    hidden_size = 1024
    num_layers = 1
    num_classes = 2

    epochs = 100
    #seq_length = 64
    learning_rate = 0.002

    word2vec = Word2Vec.load('./data/'+p+'/token/node_w2v_128').wv
    MAX_TOKENS = word2vec.syn0.shape[0]
    EMBEDDING_DIM = word2vec.syn0.shape[1]
    embeddings = np.zeros((MAX_TOKENS + 1, EMBEDDING_DIM), dtype="float32")
    embeddings[:word2vec.syn0.shape[0]] = word2vec.syn0


    model = RCF(embeddings,MAX_TOKENS+1).to(device)

    # Loss and optimizer
    l = train_dataset.label
    class_weight = 'balanced'
    classes = np.array([0,1])
    weight = compute_class_weight(class_weight,classes,l)
    criterion = nn.CrossEntropyLoss(weight = torch.from_numpy(weight).float().cuda())
    #criterion = nn.CrossEntropyLoss()
    #optimizer = torch.optim.Adam(model.parameters(), lr=learning_rate)
    optimizer = torch.optim.Adamax(model.parameters())

    tm = 0
    ot = []
    # Train the model
    total_step = len(train_loader)
    for epoch in range(epochs):
        # Set initial hidden and cell states
        model.train()
        start_time = time.time()
        for _,data in enumerate(train_loader):
            #print("--------------------")
            old,new,label = data
            label = label.to(device)
            old = torch.tensor([item.cpu().detach().numpy() for item in old]).to(device).permute(1,0).reshape(-1,50,100)
            new = torch.tensor([item.cpu().detach().numpy() for item in new]).to(device).permute(1,0).reshape(-1,50,100)
            # Forward pass
            outputs = model(old,new)
            loss = criterion(outputs,label)
            
            # Backward and optimize
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            
            if (_+1) % 10 == 0:
                end_time = time.time()
                tm += end_time-start_time
                print ('Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}, Time : {}' 
                    .format(epoch+1, epochs, _+1, total_step, loss.item(), end_time-start_time))
                start_time = time.time()
    
        print("testing",p)
        # Test the model
        model.eval()
        lb = torch.Tensor()
        pr = torch.Tensor()

        with torch.no_grad():
            correct = 0
            total = 0
            ts = len(test_loader)
            for _,data in enumerate(test_loader):
                old,new,label = data
                label = label.to(device)
                old = torch.tensor([item.cpu().detach().numpy() for item in old]).to(device).permute(1,0).reshape(-1,50,100)
                new = torch.tensor([item.cpu().detach().numpy() for item in new]).to(device).permute(1,0).reshape(-1,50,100)
                outputs = model(old,new)
                loss = criterion(outputs,label)
                __, predicted = torch.max(outputs.data, 1)
                total += label.size(0)
                pr = torch.cat((pr,predicted.cpu()),0)
                lb = torch.cat((lb,label.cpu()),0)
                correct += (predicted == label).sum().item()
                print('step :',_,' , total step :',ts,' , loss :',loss)
            print('Test Accuracy of the model on the {} test case: {} %'.format(total,100 * correct / total)) 
        
        print(pr)
        print(lb)
        zero = 0
        zero_all = 0
        one = 0
        one_all = 0
        for i in range(len(lb)):
            if lb[i]==0:
                zero_all+=1
                if pr[i]==0:zero+=1
            else:
                one_all+=1
                if pr[i]==1:one+=1
        print("Test one acc: {}/{}, zero acc: {}/{}".format(one,one_all,zero,zero_all))
        print("Recall :",metrics.recall_score(lb,pr))
        print("F1 :",metrics.f1_score(lb,pr))
        print("AUC :",metrics.roc_auc_score(lb,pr))
        print("MCC : {}".format(matthews_corrcoef(lb,pr)))
        ot.append(['','',metrics.recall_score(lb,pr),metrics.f1_score(lb,pr),metrics.roc_auc_score(lb,pr),100*correct/total,matthews_corrcoef(lb,pr),tm])
    return ot
        
if __name__ == '__main__':
    #p = 'beam'
    #nd(p)
    project = sys.argv[1]
    out = nd(project)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'project'
    ws['B1'] = 'model'
    ws['C1'] = 'Recall'
    ws['D1'] = 'F1'
    ws['E1'] = 'AUC'
    ws['F1'] = 'ACCURACY'
    ws['G1'] = 'MCC'
    ws['H1'] = 'Time'

    file_path = './rerun/DACE_'+project+'.xlsx'

    out[0][0]=project
    out[0][1]='DACE'
    for row in out:
        ws.append(row)
    wb.save(file_path)
