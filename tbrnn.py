import torch 
import torch.nn as nn
import torchvision
import torchvision.transforms as transforms
from torch.utils.data import Dataset, DataLoader
from gensim.models.word2vec import Word2Vec
import json
import random
import gensim
import numpy as np
import pandas as pd
import os,time,sys
from sklearn import metrics
import warnings
import torch.nn.utils.rnn as rnn_utils
import torch.nn.functional as F
from sklearn.utils.class_weight import compute_class_weight 
from sklearn.metrics import matthews_corrcoef
from openpyxl import Workbook,load_workbook
warnings.filterwarnings('ignore')

os.environ["CUDA_VISIBLE_DEVICES"] = "1"
# Device configuration
#device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
device = torch.device('cuda')

#Dataset
class MyDataset(Dataset):
    def __init__(self,file_path):
        labels = []
        old = []
        new = []
        source = pd.read_pickle(file_path)
        self.len = len(source)
        self.label = source['label'].tolist()
        self.old = source['old'].tolist()
        self.new = source['new'].tolist()

    def __len__(self):
        return self.len

    def __getitem__(self,index):
        return self.old[index],self.new[index],self.label[index]

class RCF(nn.Module):
    def __init__(self,weights,vocab_size):
        super(RCF,self).__init__()
        self.embedding_size = 128
        self.hidden_size = 300
        self.num_classes = 2
        self.num_layers = 1
        self.embedding=nn.Embedding(vocab_size,self.embedding_size)
        self.embedding.weight.data.copy_(torch.from_numpy(weights))
        self.lstm = nn.GRU(self.embedding_size, self.hidden_size, self.num_layers)        
        self.maxpool2 = nn.MaxPool1d(kernel_size = 30, stride = 30)
        #self.fc = nn.Linear(in_features = 600,out_features = 256)
        self.fc2 = nn.Linear(in_features = 300,out_features = 100)
        self.fc3 = nn.Linear(in_features = 100,out_features = self.num_classes)

    
    def forward(self, old, new):
        #print(old.shape)
        emb_old = self.embedding(old)
        emb_new = self.embedding(new)  
        #print(emb_old.shape)
        lstm_old,_ = self.lstm(emb_old)
        lstm_new,_ = self.lstm(emb_new)
        lstm_old = lstm_old.permute(1,2,0)
        lstm_new = lstm_new.permute(1,2,0)
        #print(lstm_old.shape)
        od = F.max_pool1d(lstm_old,lstm_old.size(2)).squeeze(2)
        on = F.max_pool1d(lstm_new,lstm_new.size(2)).squeeze(2)
        #print(od.shape)
        abs_dist = torch.abs(torch.add(od,-on))
        #print(abs_dist.shape)
        out = self.fc2(abs_dist)
        out = self.fc3(out)
        #out = F.softmax(out,dim = 1)
        return out

    def initHidden(self):
        return torch.zeros(1,1,self.hidden_size,device=device)

def nd(p):
    ss=[]
    train_path = './data/'+p+'/token/'+p+'_token_train.pkl'
    test_path = './data/'+p+'/token/'+p+'_token_test.pkl'
    valid_path = './data/'+p+'/token/'+p+'_token_dev.pkl'
    train_dataset = MyDataset(train_path)
    test_dataset = MyDataset(test_path)
    valid_dataset = MyDataset(valid_path)

    batch_size = 128

    train_loader = DataLoader(dataset = train_dataset,batch_size = batch_size,shuffle = True)
    test_loader = DataLoader(dataset = test_dataset,batch_size = batch_size,shuffle = False)
    valid_loader = DataLoader(dataset = valid_dataset,batch_size = batch_size,shuffle = False)

    input_size = 128
    hidden_size = 1024
    num_layers = 1
    num_classes = 2

    epochs = 100
    #seq_length = 64
    learning_rate = 0.002

    word2vec = Word2Vec.load('./data/'+p+'/token/node_w2v_128').wv
    MAX_TOKENS = word2vec.syn0.shape[0]
    EMBEDDING_DIM = word2vec.syn0.shape[1]
    embeddings = np.zeros((MAX_TOKENS + 1, EMBEDDING_DIM), dtype="float32")
    embeddings[:word2vec.syn0.shape[0]] = word2vec.syn0


    model = RCF(embeddings,MAX_TOKENS+1).to(device)

    # Loss and optimizer
    l = train_dataset.label
    class_weight = 'balanced'
    classes = np.array([0,1])
    weight = compute_class_weight(class_weight,classes,l)
    criterion = nn.CrossEntropyLoss(weight = torch.from_numpy(weight).float().cuda())
    #criterion = nn.CrossEntropyLoss()
    #optimizer = torch.optim.Adam(model.parameters(), lr=learning_rate)
    optimizer = torch.optim.Adamax(model.parameters())

    tt=0
    ta=0
    tb=0
    tc=0
    tm = 0
    ot = []
    # Train the model
    total_step = len(train_loader)
    for epoch in range(epochs):
        # Set initial hidden and cell states
        model.train()
        start_time = time.time()
        for _,data in enumerate(train_loader):
            #print("--------------------")
            sta = time.time()
            old,new,label = data
            label = label.to(device)
            old = torch.tensor([item.cpu().detach().numpy() for item in old]).to(device)
            new = torch.tensor([item.cpu().detach().numpy() for item in new]).to(device)
            stb = time.time()
            ta+=(stb-sta)
            # Forward pass
            outputs = model(old,new)
            loss = criterion(outputs,label)
            stc=time.time()
            tb+=(stc-stb)
            # Backward and optimize
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            
            if (_+1) % 10 == 0:
                end_time = time.time()
                tm += end_time-start_time
                print ('Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}, Time : {}' 
                    .format(epoch+1, epochs, _+1, total_step, loss.item(), end_time-start_time))
                start_time = time.time()
            std=time.time()
            tc+=(std-stc)
            tt+=(std-sta)

        print("Testing",p)
        # Test the model
        model.eval()
        lb = torch.Tensor()
        pr = torch.Tensor()

        with torch.no_grad():
            correct = 0
            total = 0
            ts = len(test_loader)
            for _,data in enumerate(test_loader):
                old,new,label = data
                label = label.to(device)
                old = torch.tensor([item.cpu().detach().numpy() for item in old]).to(device)
                new = torch.tensor([item.cpu().detach().numpy() for item in new]).to(device)
                outputs = model(old,new)
                loss = criterion(outputs,label)
                __, predicted = torch.max(outputs.data, 1)
                total += label.size(0)
                pr = torch.cat((pr,predicted.cpu()),0)
                lb = torch.cat((lb,label.cpu()),0)
                correct += (predicted == label).sum().item()
                print('step :',_,' , total step :',ts,' , loss :',loss)
            print('Test Accuracy of the model on the {} test case: {} %'.format(total,100 * correct / total)) 
        
        print(pr)
        print(lb)
        zero = 0
        zero_all = 0
        one = 0
        one_all = 0
        for i in range(len(lb)):
            if lb[i]==0:
                zero_all+=1
                if pr[i]==0:zero+=1
            else:
                one_all+=1
                if pr[i]==1:one+=1
        print("Test one acc: {}/{}, zero acc: {}/{}".format(one,one_all,zero,zero_all))
        print("Recall :",metrics.recall_score(lb,pr))
        print("F1 :",metrics.f1_score(lb,pr))
        print("AUC :",metrics.roc_auc_score(lb,pr))
        print("MCC :",matthews_corrcoef(lb,pr))
        ot.append(['','',metrics.recall_score(lb,pr),metrics.f1_score(lb,pr),metrics.roc_auc_score(lb,pr),100*correct/total,matthews_corrcoef(lb,pr),tm]
)
    return ot
        #return metrics.recall_score(lb,pr),metrics.f1_score(lb,pr),metrics.roc_auc_score(lb,pr),100 * correct / total,tt,ta,tb,tc

if __name__ == '__main__':
    #p = 'beam'
    #nd(p)
    project = sys.argv[1]
    out = nd(project)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'project'
    ws['B1'] = 'model'
    ws['C1'] = 'Recall'
    ws['D1'] = 'F1'
    ws['E1'] = 'AUC'
    ws['F1'] = 'ACCURACY'
    ws['G1'] = 'MCC'
    ws['H1'] = 'Time'

    file_path = './rerun/TBRNN_'+project+'.xlsx'

    out[0][0]=project
    out[0][1]='TBRNN'
    for row in out:
        ws.append(row)
    wb.save(file_path)
