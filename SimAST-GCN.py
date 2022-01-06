#GAT_Concate
import torch 
import torch.nn as nn
import torchvision
import torchvision.transforms as transforms
from torch.utils.data import Dataset, DataLoader,WeightedRandomSampler
from gensim.models.word2vec import Word2Vec
import json
import random
import gensim
import numpy as np
import pandas as pd
import os,time
import warnings
import torch.nn.functional as F
import torch.nn.utils.rnn as rnn_utils
from sklearn import metrics
from sklearn.utils.class_weight import compute_class_weight 
from sklearn.metrics import matthews_corrcoef
import logging
import sys
from openpyxl import Workbook,load_workbook

from data_iter import MyClassBalanceDataset,MyBatchSampler,MyDataset

warnings.filterwarnings('ignore')


os.environ["CUDA_VISIBLE_DEVICES"] = "0"
# Device configuration
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
#device = torch.device('cpu')


#Module
class GraphConvolution(nn.Module):
    def __init__(self, in_features, out_features, bias=True):
        super(GraphConvolution, self).__init__()
        self.in_features = in_features
        self.out_features = out_features
        self.weight = nn.Parameter(torch.FloatTensor(in_features, out_features))
        #nn.init.xavier_uniform_(self.weight.data, gain=1.414)
        torch.nn.init.kaiming_uniform_(self.weight.data)
        if bias:
            self.bias = nn.Parameter(torch.FloatTensor(out_features))
            #torch.nn.init.kaiming_uniform_(self.bias.data)
            #nn.init.xavier_uniform_(self.bias.data, gain=1.414)
        else:
            self.register_parameter('bias', None)
        self.reset_parameter()
        self.batch_norm = nn.BatchNorm1d(1000) #fixed, corresponding to the max len of the token

    def reset_parameter(self):
        torch.nn.init.kaiming_uniform_(self.weight, a = math.sqrt(5))
        if self.bias is not None:
            fan_in, _ = torch.nn.init._calculate_fan_in_and_fan_out(self.weight)
            bound = 1/math.sqrt(fan_in)
            torch.nn.init.uniform_(self.bias, -bound, bound)

    def forward(self, text, adj):
        #print(text.shape,adj.shape,self.weight.shape)
        hidden = torch.matmul(text, self.weight)
        denom = torch.sum(adj, dim=2, keepdim=True) + 1
        #print(hidden.shape,denom.shape)
        output = torch.matmul(adj, hidden)
        output = output / denom
        if self.bias is not None:
            return self.batch_norm(output + self.bias)
        else:
            return self.batch_norm(output)

class GC(nn.Module):
    def __init__(self, in_features, out_features, bias=True):
        super(GC, self).__init__()
        self.in_features = in_features
        self.out_features = out_features
        self.W_at = nn.Linear(in_features, out_features)
        self.U_at_1 = nn.Linear(in_features, out_features)

    def forward(self, text, adj):
        denom = adj.sum(2).unsqueeze(2)+1
        att = adj.bmm(text)
        att = att / denom
        forg = torch.sigmoid(self.W_at(att)+self.U_at_1(text))


class CIAN(nn.Module):
    def __init__(self, weights, vocab_size, gcnn):
        super(CIAN, self).__init__()
        self.embedding_size = 300
        self.hidden_size = 300
        self.embedding=nn.Embedding(vocab_size,self.embedding_size)
        #print('weights: ',weights)
        self.embedding.weight.data.copy_(weights)
        #self.embedding = nn.Embedding.from_pretrained(weights,freeze=True)
        self.bigru1 = nn.GRU(self.embedding_size,self.hidden_size,num_layers=1,bidirectional=True,batch_first=True)  #,batch_first=True  ,batch_first=True
        #self.bigru2 = nn.GRU(self.embedding_size,self.hidden_size,num_layers=1,bidirectional=True,batch_first=True)

        self.gc1 = nn.ModuleList([GraphConvolution(2*self.hidden_size,2*self.hidden_size) for i in range(gcnn)])

        #self.gc2 = nn.ModuleList([GraphConvolution(2*self.hidden_size,2*self.hidden_size) for i in range(gcnn)])

        #self.gc1 = GraphConvolution(2*self.hidden_size, 2*self.hidden_size)
        #self.gc2 = GraphConvolution(2*self.hidden_size, 2*self.hidden_size)
        #self.gc3 = GraphConvolution(2*self.hidden_size, 2*self.hidden_size)
        #self.gc4 = GraphConvolution(2*self.hidden_size, 2*self.hidden_size)
        self.fc1 = nn.Linear(in_features = 2*self.hidden_size,out_features = self.hidden_size)
        self.fc2 = nn.Linear(in_features = self.hidden_size,out_features = 150)
        self.fc3 = nn.Linear(in_features = 150,out_features = 2)
        self.dropout = nn.Dropout(0.3)
    
    def forward(self, old, go, new, gn):
        #print(old.shape)
        eo = self.embedding(old)
        en = self.embedding(new)

        #eo = self.dropout(eo)
        #en = self.dropout(en)
        oo,_ = self.bigru1(eo)
        on,_ = self.bigru1(en)

        #oo = oo.permute(1,0,2)
        #on = on.permute(1,0,2)
        #print(oo.shape)

        o = oo
        for gcn in self.gc1:
            o = F.leaky_relu(gcn(o,go))

        n = on
        for gcn in self.gc1:
            n = F.leaky_relu(gcn(n,go))

        alpha_mat = torch.matmul(o, oo.transpose(1, 2))
        alpha = F.softmax(alpha_mat.sum(1, keepdim=True), dim=2)

        beta_mat = torch.matmul(n, on.transpose(1, 2))
        beta = F.softmax(beta_mat.sum(1, keepdim=True), dim=2)

        o = torch.matmul(alpha, oo).squeeze(1)
        #print(o.shape)
        n = torch.matmul(beta, on).squeeze(1)
        #o = F.normalize(o, p=2, dim=1)
        #n = F.normalize(n, p=2, dim=1)
        abs_dist = torch.abs(torch.add(o,-n))
        #abs_dist = torch.cat([o,n],1)
        #print(abs_dist)
        ot = F.relu(self.fc1(abs_dist))
        ot = F.relu(self.fc2(ot))
        ot = self.fc3(ot)
        #print(ot)
        '''
        oo = oo.permute(0,2,1)
        on = on.permute(0,2,1)
        l_att = F.softmax(torch.tanh(torch.einsum('ijk,kl,ilm->ijm',oo,self.l_w,torch.unsqueeze(new_avg,-1))+self.l_b),dim=1)
        r_att = F.softmax(torch.tanh(torch.einsum('ijk,kl,ilm->ijm',on,self.r_w,torch.unsqueeze(old_avg,-1))+self.r_b),dim=1)
        l_req = torch.sum(l_att*oo,1)
        r_req = torch.sum(r_att*on,1)
        req = torch.cat((l_req,r_req),1)
        ot = self.fc1(req)
        ot = self.fc2(ot)
        '''
        #ot = F.softmax(ot,dim = 1)
        return ot
        

def cian(p,gcnn):
    ot = []
    train_path = './data/'+p+'/train.pkl'
    test_path = './data/'+p+'/test.pkl'
    valid_path = './data/'+p+'/dev.pkl'
    #train_dataset = MyClassBalanceDataset(train_path)
    train_dataset = MyDataset(train_path)
    test_dataset = MyDataset(test_path)
    #valid_dataset = MyDataset(valid_path)

    #over sample
    '''
    train_rate = np.bincount(train_dataset.label)
    class_num = train_rate.tolist()
    train_weight = 1./torch.tensor(class_num,dtype =torch.float)
    train_sample = train_weight[train_dataset.label]
    train_sp = WeightedRandomSampler(weights = train_sample, num_samples = len(train_sample))
'''

    batch_size = 128

    #tr = np.bincount(train_dataset.label)
    #cn = tr.tolist()
    #cc = [cn[0]/sum(cn),cn[1]/sum(cn)]
    #batchSampler = MyBatchSampler(train_dataset, batch_size, cc)

    #train_loader = DataLoader(train_dataset,batch_sampler = batchSampler)
    train_loader = DataLoader(dataset = train_dataset,batch_size = batch_size,shuffle = True)
    #train_loader = DataLoader(dataset = train_dataset,sampler = train_sp, batch_size = batch_size, shuffle = False)  #over sample
    #train_loader = DataLoader(dataset = train_dataset,batch_size = batch_size,shuffle = True)
    test_loader = DataLoader(dataset = test_dataset,batch_size = batch_size,shuffle = False)
    #valid_loader = DataLoader(dataset = valid_dataset,batch_size = batch_size,shuffle = False)

    epochs = 100
    #learning_rate = 0.0002

    word2vec = Word2Vec.load('./data/'+p+'/node_w2v_128').wv
    MAX_TOKENS = word2vec.syn0.shape[0]
    EMBEDDING_DIM = word2vec.syn0.shape[1]
    embeddings = np.zeros((MAX_TOKENS + 1, EMBEDDING_DIM), dtype="float32")
    embeddings[:word2vec.syn0.shape[0]] = word2vec.syn0
    embeddings = torch.tensor(embeddings).to(device)

    model = CIAN(embeddings,MAX_TOKENS+1,gcnn).to(device)
    #def weights_init(m):                                               
    #    nn.init.normal_(m.weight.data, 0.0, 0.02)  
    def weights_init(m):                                        
        classname = m.__class__.__name__                         
        if classname.find('Linear') != -1:                          
            nn.init.normal_(m.weight, 0.0, 0.02)              
                        
    model.apply(weights_init)   
    #print(model.parameters())

    # Loss and optimizer
    
    l = train_dataset.label
    class_weight = 'balanced'
    classes = np.array([0,1])
    weight = compute_class_weight(class_weight = class_weight,classes = classes, y = l)
    criterion = nn.CrossEntropyLoss(weight = torch.from_numpy(weight).float().cuda())  #
    
    #criterion = nn.CrossEntropyLoss()
    #optimizer = torch.optim.Adamax(model.parameters()) #, lr=learning_rate  ,weight_decay=0.00001
    optimizer = torch.optim.Adam(model.parameters(),lr=0.001,weight_decay=0.00001)#weight_decay=0.00001  #adamax 0.0002
    #optimizer = torch.optim.SGD(model.parameters(), lr=0.01, momentum=0.9, weight_decay=0.00001)

    #cal time
    tm = 0
    ta=0
    tb=0
    tc=0
    # Train the model
    total_step = len(train_loader)
    for epoch in range(epochs):
        logging.info("train epoch: "+str(epoch))
        #("Training ",p)
        model.train()
        stm = time.time()
        start_time = time.time()
        ss = time.time()
        #if epoch == 10: model.embedding.weight.requires_grad = True
        for _,data in enumerate(train_loader):
            stz = time.time()
            old,go,new,gn,label = data
            #print(old[0],new[0])
            #print(old[1],new[1])
            #return []
            #print(old,go,new,gn,label)
            label = label.to(device)
            go = go.to(device).float()
            gn = gn.to(device).float()
            old = old.to(device).int()
            new = new.to(device).int()
            #new = torch.tensor([item.cpu().detach().numpy() for item in new]).to(device).int()
            sta = time.time()
            ta+=(sta-stz)#data load time
            # Forward pass
            outputs = model(old,go,new,gn)
            #print(outputs.cpu())
            loss = criterion(outputs,label)
            stb = time.time()
            tb+=(stb-sta)#forward time
            # Backward and optimize
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            stc = time.time()
            tc+=(stc-stb)
            
            if (_+1) % 10 == 0:
                end_time = time.time()
                logging.info('Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}, Time: {}' 
                    .format(epoch+1, epochs, _+1, total_step, loss.item(), end_time - start_time))
                #print ('Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}, Time: {}' 
                #    .format(epoch+1, epochs, _+1, total_step, loss.item(), end_time - start_time))
                start_time = time.time()
        logging.info("finish training eopoch : "+str(epoch))
        etm = time.time()
        ee = time.time()
        tm += (ee-ss)
        #print("timing : ",tm,epoch)
        logging.info("timing : {}, {}".format(tm,epoch))
        model.eval()
        logging.info("Validating : {}".format(p))
        #print('Validating ',p)
        lb = torch.Tensor()
        pr = torch.Tensor()
        with torch.no_grad():
            correct = 0
            total = 0
            ts = len(test_loader)
            for _,data in enumerate(test_loader):
                old,go,new,gn,label = data
                label = label.to(device)
                go = go.to(device).float()
                gn = gn.to(device).float()
                old = old.to(device).int()
                new = new.to(device).int()
                outputs = model(old,go,new,gn)
                loss = criterion(outputs,label)
                __, predicted = torch.max(outputs.data, 1)
                total += label.size(0)
                pr = torch.cat((pr,predicted.cpu()),0)
                lb = torch.cat((lb,label.cpu()),0)
                correct += (predicted == label).sum().item()
                #print('step :',_,' , total step :',ts,' , loss :',loss)
            #print('Test Accuracy of the model on the {} test case: {} %'.format(total,100 * correct / total)) 
        #print(pr)
        #print(lb)
        zero = 0
        zero_all = 0
        one = 0
        one_all = 0
        for i in range(len(lb)):
            if lb[i]==0:
                zero_all+=1
                if pr[i]==0:zero+=1
            else:
                one_all+=1
                if pr[i]==1:one+=1
        logging.info("Test one acc: {}/{}, zero acc: {}/{}".format(one,one_all,zero,zero_all))
        logging.info("Recall : {}".format(metrics.recall_score(lb,pr)))
        logging.info("F1 : {}".format(metrics.f1_score(lb,pr)))
        logging.info("AUC : {}".format(metrics.roc_auc_score(lb,pr)))
        logging.info("MCC : {}".format(matthews_corrcoef(lb,pr)))
        if epoch>=0: ot.append(['','',metrics.recall_score(lb,pr),metrics.f1_score(lb,pr),metrics.roc_auc_score(lb,pr),100*correct/total,matthews_corrcoef(lb,pr),tm])
    return ot
    '''
        start_time = time.time()
        for _,data in enumerate(valid_loader):
            old,new,label = data
            label = label.to(device)
            old = torch.tensor([item.cpu().detach().numpy() for item in old]).to(device)
            new = torch.tensor([item.cpu().detach().numpy() for item in new]).to(device)
            outputs = model(old,new)
            loss = criterion(outputs,label)      
            if (_+1) % 10 == 0:
                end_time = time.time()
                print ('Valid, Step [{}/{}], Loss: {:.4f}, Time: {}' 
                    .format(_+1, len(valid_loader), loss.item(), end_time - start_time))
                start_time = time.time()
                '''

'''
    print("Testing ",p)
    model.eval()
    lb = torch.Tensor()
    pr = torch.Tensor()

    with torch.no_grad():
        correct = 0
        total = 0
        ts = len(test_loader)
        for _,data in enumerate(test_loader):
            old,new,label = data
            label = label.to(device)
            #old = old.to(device)
            #new = new.to(device)
            old = torch.tensor([item.cpu().detach().numpy() for item in old]).to(device)
            new = torch.tensor([item.cpu().detach().numpy() for item in new]).to(device)
            outputs = model(old,new)
            loss = criterion(outputs,label)
            __, predicted = torch.max(outputs.data, 1)
            total += label.size(0)
            pr = torch.cat((pr,predicted.cpu()),0)
            lb = torch.cat((lb,label.cpu()),0)
            correct += (predicted == label).sum().item()
            print('step :',_,' , total step :',ts,' , loss :',loss)
        print('Test Accuracy of the model on the {} test case: {} %'.format(total,100 * correct / total)) 
    
    print(pr)
    print(lb)
    zero = 0
    zero_all = 0
    one = 0
    one_all = 0
    for i in range(len(lb)):
        if lb[i]==0:
            zero_all+=1
            if pr[i]==0:zero+=1
        else:
            one_all+=1
            if pr[i]==1:one+=1
    print("Test one acc: {}/{}, zero acc: {}/{}".format(one,one_all,zero,zero_all))
    print("Recall :",metrics.recall_score(lb,pr))
    print("F1 :",metrics.f1_score(lb,pr))
    print("AUC :",metrics.roc_auc_score(lb,pr))
    return metrics.recall_score(lb,pr),metrics.f1_score(lb,pr),metrics.roc_auc_score(lb,pr),100*correct/total,tm,ta,tb,tc,matthews_corrcoef(lb,pr)
'''
if __name__ == '__main__':
    project = sys.argv[1]
    gcnn = sys.argv[2]

    logging.basicConfig(level=logging.INFO,
                        filename='./rerun/SimASTGCN_'+project+'_'+gcnn+'.log',
                        filemode='a',
                        format='%(asctime)s %(levelname)s %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S')
    #p = 'accumulo'
    out = cian(project,int(gcnn))

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'project'
    ws['B1'] = 'model'
    ws['C1'] = 'Recall'
    ws['D1'] = 'F1'
    ws['E1'] = 'AUC'
    ws['F1'] = 'ACCURACY'
    ws['G1'] = 'MCC'
    ws['H1'] = 'Time'

    file_path = './rerun/SimASTGCN_'+project+'_'+gcnn+'.xlsx'

    out[0][0]=project
    out[0][1]='SimASTGCN'
    for row in out:
        ws.append(row)
    wb.save(file_path)
